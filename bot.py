from telegram import Update
from telegram.ext import Application, CommandHandler, MessageHandler, filters, CallbackContext
from openpyxl import load_workbook

def load_products(file_name):
    products = []
    wb = load_workbook(file_name)
    for sheet_name in ['Table 1', 'Table 2', 'Table 3', 'Table 4', 'Table 5']:  # نام شیت‌های صحیح با فاصله
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) >= 4:
                code, product_name, car_name, price = str(row[0]), str(row[1]), str(row[2]), str(row[3])
                product_info = {
                    'code': code,
                    'name': product_name,
                    'car': car_name,
                    'price': price,
                    'sheet': sheet_name,
                    'row': sheet._current_row
                }
                products.append(product_info)
    return products

# لیست محصولات
file_name = 'List.xlsx'
products = load_products(file_name)

async def start(update: Update, context: CallbackContext) -> None:
    await update.message.reply_text('ربات AguPart  جهت جستجو محصولات و، اتحاد طلایی آسیا')

async def search(update: Update, context: CallbackContext) -> None:
    query = update.message.text.lower()
    result = [f"شیت: {product['sheet']}, ردیف: {product['row']}, قیمت : {product['code']}, خودرو : {product['name']}, نام محصول : {product['car']}, کد کالا: {product['price']}"
              for product in products if query in product['name'].lower() or query in product['car'].lower() or query in product['code'].lower() or query in product['price'].lower()]
    if result:
        await update.message.reply_text("\n\n".join(result))
    else:
        await update.message.reply_text('محصولی با این نام یافت نشد.')

def main():
    application = Application.builder().token("7193773661:AAHO9LkqKZSVzNShKhTq4h0Y5xk01B1patI").build()

    application.add_handler(CommandHandler("start", start))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, search))

    application.run_polling()

if __name__ == '__main__':
    main()